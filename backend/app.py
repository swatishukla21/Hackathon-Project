from flask import Flask, request, jsonify
from flask_cors import CORS

import openpyxl
import re
import os
import networkx as nx

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ---------------- Existing Upload Endpoint ----------------
@app.route("/upload", methods=["POST"])
def upload_file():
    global last_graph

    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    path = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(path)

    wb = openpyxl.load_workbook(path, data_only=False)

    node_set = set()
    links = []
    formulas = []

    same_sheet_ref = r"[A-Za-z]+[0-9]+"
    cross_sheet_ref = r"([A-Za-z0-9_]+)!([A-Za-z]+[0-9]+)"

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                cell_id = f"{sheet_name}!{cell.coordinate}"
                node_set.add(cell_id)

                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({
                        "id": cell_id,
                        "formula": cell.value
                    })

                    for ref in re.findall(same_sheet_ref, cell.value):
                        ref_id = f"{sheet_name}!{ref}"
                        node_set.add(ref_id)
                        links.append({"source": ref_id, "target": cell_id})

                    for other_sheet, ref in re.findall(cross_sheet_ref, cell.value):
                        ref_id = f"{other_sheet}!{ref}"
                        node_set.add(ref_id)
                        links.append({"source": ref_id, "target": cell_id})

    nodes = [{"id": n} for n in node_set]

    # -------- store graph in memory for dependency API --------
    G = nx.DiGraph()
    G.add_nodes_from(n for n in node_set)
    for l in links:
        G.add_edge(l["source"], l["target"])

    last_graph = G  # save globally

    return jsonify({
        "nodes": nodes,
        "links": links,
        "formulas": formulas
    })


# ---------------- NEW ENDPOINT: dependency chain ----------------
@app.route("/dependency_chain", methods=["POST"])
def dependency_chain():
    global last_graph

    data = request.get_json()
    node = data.get("node")

    if "last_graph" not in globals():
        return jsonify({"error": "No graph generated yet"}), 400

    if node not in last_graph.nodes:
        return jsonify({"error": "Node not found"}), 404

    # upstream = all nodes that lead into this node
    upstream = nx.ancestors(last_graph, node)

    # downstream = all nodes this node feeds into
    downstream = nx.descendants(last_graph, node)

    return jsonify({
        "upstream": list(upstream),
        "downstream": list(downstream)
    })


if __name__ == "__main__":
    app.run(debug=True)
